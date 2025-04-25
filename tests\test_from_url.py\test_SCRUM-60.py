import os
import shutil
import tempfile
import pytest
from unittest.mock import patch
from extractor import RequirementExtractor, ExtractionError

def create_sample_pdf(path):
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', size=12)
    pdf.cell(200, 10, txt='Requirement: The system shall support PDF extraction.', ln=True)
    pdf.output(path)

def create_sample_docx(path):
    from docx import Document
    doc = Document()
    doc.add_paragraph('Requirement: The system shall support Word document extraction.')
    doc.save(path)

def create_sample_xlsx(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Requirement'
    ws['A2'] = 'The system shall support Excel extraction.'
    wb.save(path)

def create_sample_image(path):
    from PIL import Image, ImageDraw, ImageFont
    img = Image.new('RGB', (400, 100), color=(255, 255, 255))
    d = ImageDraw.Draw(img)
    d.text((10, 40), 'Requirement: The system shall support image extraction.', fill=(0, 0, 0))
    img.save(path)

@pytest.fixture(scope='module')
def temp_dir():
    dirpath = tempfile.mkdtemp()
    yield dirpath
    shutil.rmtree(dirpath)

@pytest.fixture(scope='module')
def extractor():
    return RequirementExtractor()

def test_extract_from_pdf(temp_dir, extractor):
    pdf_path = os.path.join(temp_dir, 'sample.pdf')
    create_sample_pdf(pdf_path)
    requirements = extractor.extract(pdf_path)
    assert isinstance(requirements, list)
    assert any('PDF extraction' in req for req in requirements)

def test_extract_from_docx(temp_dir, extractor):
    docx_path = os.path.join(temp_dir, 'sample.docx')
    create_sample_docx(docx_path)
    requirements = extractor.extract(docx_path)
    assert isinstance(requirements, list)
    assert any('Word document extraction' in req for req in requirements)

def test_extract_from_xlsx(temp_dir, extractor):
    xlsx_path = os.path.join(temp_dir, 'sample.xlsx')
    create_sample_xlsx(xlsx_path)
    requirements = extractor.extract(xlsx_path)
    assert isinstance(requirements, list)
    assert any('Excel extraction' in req for req in requirements)

def test_extract_from_image(temp_dir, extractor):
    image_path = os.path.join(temp_dir, 'sample.png')
    create_sample_image(image_path)
    requirements = extractor.extract(image_path)
    assert isinstance(requirements, list)
    assert any('image extraction' in req for req in requirements)

def test_extract_from_unsupported_format(temp_dir, extractor):
    txt_path = os.path.join(temp_dir, 'sample.txt')
    with open(txt_path, 'w') as f:
        f.write('Requirement: This is a plain text file.')
    with pytest.raises(ExtractionError):
        extractor.extract(txt_path)

def test_extract_from_corrupted_pdf(temp_dir, extractor):
    pdf_path = os.path.join(temp_dir, 'corrupted.pdf')
    with open(pdf_path, 'wb') as f:
        f.write(b'%PDF-1.4 corrupted content')
    with pytest.raises(ExtractionError):
        extractor.extract(pdf_path)

def test_extract_from_empty_file(temp_dir, extractor):
    empty_path = os.path.join(temp_dir, 'empty.pdf')
    open(empty_path, 'wb').close()
    with pytest.raises(ExtractionError):
        extractor.extract(empty_path)

def test_extract_with_ai_failure(temp_dir, extractor):
    pdf_path = os.path.join(temp_dir, 'sample_ai_fail.pdf')
    create_sample_pdf(pdf_path)
    with patch.object(extractor, 'ai_model_extract', side_effect=ExtractionError('AI failure')):
        with pytest.raises(ExtractionError):
            extractor.extract(pdf_path)

def test_extract_structure_of_output(temp_dir, extractor):
    docx_path = os.path.join(temp_dir, 'sample_structure.docx')
    create_sample_docx(docx_path)
    requirements = extractor.extract(docx_path)
    assert isinstance(requirements, list)
    for req in requirements:
        assert isinstance(req, str)
        assert req.strip() != ''
